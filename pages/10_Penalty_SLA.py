import streamlit as st
import pandas as pd
import plotly.express as px
import sys
import os
from io import BytesIO
from datetime import datetime

# Openpyxl for Professional Excel Formatting
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from utils.data_processing import filter_by_period
from utils.auto_load import auto_load_tickets

st.set_page_config(page_title="Penalty & SLA | XTRNATE", page_icon="📜", layout="wide")

# Exact CKT Page Custom CSS Theme
st.markdown("""
<style>
@media (max-width: 768px) {
  .block-container { padding: 0.6rem !important; }
}

.ckt-hero {
  background: linear-gradient(135deg, #0f172a 0%, #1e3a8a 70%);
  border: 1px solid #38bdf8;
  border-radius: 18px;
  padding: 1.4rem 1.6rem 1.1rem 1.6rem;
  margin-bottom: 1.2rem;
  box-shadow: 0 10px 30px rgba(15,23,42,0.35);
}
.ckt-hero h1 { color: #fff; margin: 0 0 0.25rem 0; font-size: 1.7rem; }
.ckt-hero p { color: #cbd5e1; margin: 0; }

.ckt-card {
  background: #0f172a;
  border: 1px solid #334155;
  border-radius: 16px;
  padding: 1.2rem 1.4rem;
  margin-top: 0.8rem;
  margin-bottom: 1rem;
}
.ckt-label { color: #94a3b8; font-size: 0.78rem; letter-spacing: 0.06em; text-transform: uppercase; margin-bottom: 0.2rem; }
</style>
""", unsafe_allow_html=True)

# Hero Banner Integration
st.markdown("""
<div class="ckt-hero">
  <h1>📜 Automated Penalty — HCIN vs ONEOTT</h1>
  <p>Period-wise SLA breach &nbsp;•&nbsp; Site-wise down count & total downtime tracking</p>
</div>
""", unsafe_allow_html=True)

if st.session_state.get('closed_df') is None:
    with st.spinner("Auto-loading data..."):
        auto_load_tickets()

if 'selected_isp' not in st.session_state:
    st.session_state.selected_isp = "ALL"

closed_df = st.session_state.get('closed_df')
open_df = st.session_state.get('open_df')

if closed_df is None or closed_df.empty:
    st.warning("Closed data nahi mila. Sheet share check karo.")
    st.stop()

period = st.radio("Period", ["Last 1 Month", "Last 3 Months", "Last 6 Months", "Overall"], horizontal=True)
period_map = {"Last 1 Month": "1M", "Last 3 Months": "3M", "Last 6 Months": "6M", "Overall": "ALL"}
df_all = filter_by_period(closed_df, period_map[period]) if period_map[period] != "ALL" else closed_df.copy()

if 'resolution_days' not in df_all.columns:
    st.error("resolution_days nahi hai. Submitted + Resolved Time-Active chahiye.")
    st.stop()

if 'penalty_rules' not in st.session_state:
    st.session_state.penalty_rules = {
        'l1_hours': 24, 'l1_penalty': 500,
        'l2_hours': 72, 'l2_penalty': 2000,
        'l3_hours': 120, 'l3_penalty': 5000,
    }

with st.expander("⚙️ Penalty Rules Configuration"):
    r = st.session_state.penalty_rules
    c1, c2, c3 = st.columns(3)
    with c1:
        r['l1_hours'] = st.number_input("L1 SLA (hrs)", value=int(r['l1_hours']), min_value=1)
        r['l1_penalty'] = st.number_input("L1 Penalty ₹", value=int(r['l1_penalty']), min_value=0, step=100)
    with c2:
        r['l2_hours'] = st.number_input("L2 SLA (hrs)", value=int(r['l2_hours']), min_value=1)
        r['l2_penalty'] = st.number_input("L2 Penalty ₹", value=int(r['l2_penalty']), min_value=0, step=100)
    with c3:
        r['l3_hours'] = st.number_input("L3 SLA (hrs)", value=int(r['l3_hours']), min_value=1)
        r['l3_penalty'] = st.number_input("L3 Penalty ₹", value=int(r['l3_penalty']), min_value=0, step=500)
    st.session_state.penalty_rules = r

rules = st.session_state.penalty_rules

def apply_penalty(df):
    if df is None or df.empty:
        return pd.DataFrame()
    d = df.copy()
    d['resolution_hours'] = pd.to_numeric(d['resolution_days'], errors='coerce') * 24

    def calc(hours):
        if pd.isna(hours):
            return 'Unknown', 0
        if hours > rules['l3_hours']:
            return 'L3 Critical', rules['l3_penalty']
        if hours > rules['l2_hours']:
            return 'L2 Breach', rules['l2_penalty']
        if hours > rules['l1_hours']:
            return 'L1 Breach', rules['l1_penalty']
        return 'Within SLA', 0

    res = d['resolution_hours'].apply(calc)
    d['sla_status'] = res.apply(lambda x: x[0])
    d['penalty_est'] = res.apply(lambda x: x[1])
    return d

def get_isp_slice(df, name):
    if df is None or df.empty or 'isp' not in df.columns:
        return pd.DataFrame()
    return df[df['isp'] == name].copy()

def site_summary(d):
    """Per site: count, avg hours, total downtime, penalty."""
    if d is None or d.empty or 'site_code' not in d.columns:
        return pd.DataFrame()
    x = d.copy()
    if 'down_time_min' not in x.columns:
        x['down_time_min'] = x.get('resolution_hours', 0) * 60
    x['down_time_min'] = pd.to_numeric(x['down_time_min'], errors='coerce').fillna(0)
    x['resolution_hours'] = pd.to_numeric(x.get('resolution_hours', 0), errors='coerce').fillna(0)
    x['penalty_est'] = pd.to_numeric(x.get('penalty_est', 0), errors='coerce').fillna(0)

    g = x.groupby('site_code').agg(
        down_count=('ticket_id', 'count'),
        total_downtime_min=('down_time_min', 'sum'),
        avg_resolution_hrs=('resolution_hours', 'mean'),
        max_resolution_hrs=('resolution_hours', 'max'),
        total_penalty_inr=('penalty_est', 'sum'),
    ).reset_index()
    g['total_downtime_hrs'] = (g['total_downtime_min'] / 60).round(1)
    g['avg_resolution_hrs'] = g['avg_resolution_hrs'].round(1)
    g['max_resolution_hrs'] = g['max_resolution_hrs'].round(1)
    g['total_penalty_inr'] = g['total_penalty_inr'].astype(int)
    g = g.sort_values(['down_count', 'total_downtime_hrs'], ascending=False)

    if 'state' in x.columns:
        g['state'] = g['site_code'].map(x.groupby('site_code')['state'].first())
    return g

hcin = apply_penalty(get_isp_slice(df_all, 'HCIN'))
ott = apply_penalty(get_isp_slice(df_all, 'ONEOTT'))

def summary_block(d):
    if d.empty:
        return {'total': 0, 'within': 0, 'l1': 0, 'l2': 0, 'l3': 0, 'penalty': 0, 'breaches': 0}
    return {
        'total': len(d),
        'within': int((d['sla_status'] == 'Within SLA').sum()),
        'l1': int((d['sla_status'] == 'L1 Breach').sum()),
        'l2': int((d['sla_status'] == 'L2 Breach').sum()),
        'l3': int((d['sla_status'] == 'L3 Critical').sum()),
        'penalty': int(d['penalty_est'].sum()),
        'breaches': int((d['penalty_est'] > 0).sum()),
    }

h = summary_block(hcin)
o = summary_block(ott)

# Summary Cards formatted inside Dark Container Theme
with st.container():
    st.subheader("⚡ HCIN vs ONEOTT Penalty Summary")
    col_h, col_o = st.columns(2)
    with col_h:
        st.markdown("""
        <div class="ckt-card">
          <div class="ckt-label">Vendor Overview</div>
          <h3 style="color:#38bdf8; margin:0;">🏢 HCIN</h3>
        </div>
        """, unsafe_allow_html=True)
        st.metric("Total Tickets", h['total'])
        a, b, c, d_ = st.columns(4)
        a.metric("Within", h['within']); b.metric("L1", h['l1']); c.metric("L2", h['l2']); d_.metric("L3", h['l3'])
        st.metric("Penalty ₹", f"{h['penalty']:,}")
    with col_o:
        st.markdown("""
        <div class="ckt-card">
          <div class="ckt-label">Vendor Overview</div>
          <h3 style="color:#f97316; margin:0;">🌐 ONEOTT</h3>
        </div>
        """, unsafe_allow_html=True)
        st.metric("Total Tickets", o['total'])
        a, b, c, d_ = st.columns(4)
        a.metric("Within", o['within']); b.metric("L1", o['l1']); c.metric("L2", o['l2']); d_.metric("L3", o['l3'])
        st.metric("Penalty ₹", f"{o['penalty']:,}")

fig = px.bar(pd.DataFrame({'ISP': ['HCIN', 'ONEOTT'], 'Penalty_INR': [h['penalty'], o['penalty']]}),
             x='ISP', y='Penalty_INR', color='ISP',
             color_discrete_map={'HCIN': '#38bdf8', 'ONEOTT': '#f97316'}, text='Penalty_INR')
fig.update_layout(template='plotly_dark', height=320)
st.plotly_chart(fig, use_container_width=True)

st.markdown("---")
st.subheader(f"📍 Site-wise Down Count & Downtime — {period}")
st.caption("Har site: kitni baar down | avg/max hours | total downtime | estimated penalty")

tab_h, tab_o, tab_all = st.tabs(["🏢 HCIN Sites", "🌐 ONEOTT Sites", "📋 Combined Data"])

h_sites = site_summary(hcin)
o_sites = site_summary(ott)

# Styled Dark Excel Exporter Function
def get_styled_excel_bytes(hcin_df, ott_df, combined_df):
    out = BytesIO()
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet

    # Fonts & Styles matching UI Dark Theme
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    data_font = Font(name=font_family, size=10, color="000000")
    site_font = Font(name=font_family, size=10, bold=True, color="1E3A8A")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    sheets_to_create = []
    if not hcin_df.empty:
        sheets_to_create.append(('HCIN_Sites', hcin_df))
    if not ott_df.empty:
        sheets_to_create.append(('ONEOTT_Sites', ott_df))
    if not combined_df.empty:
        sheets_to_create.append(('Combined_Sites', combined_df))

    for sheet_name, df_data in sheets_to_create:
        ws = wb.create_sheet(title=sheet_name)
        
        # Write Headers
        headers = list(df_data.columns)
        ws.append(headers)
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
            ws.row_dimensions[1].height = 26

        # Write Rows with Alternating Colors & Border Formatting
        for row_idx, row in enumerate(df_data.itertuples(index=False), start=2):
            ws.append(list(row))
            ws.row_dimensions[row_idx].height = 20
            is_alt = (row_idx % 2 == 0)
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.font = data_font
                
                if is_alt:
                    cell.fill = alt_row_fill
                
                header_name = str(headers[col_idx - 1]).lower()
                
                # Highlight Site Code specifically
                if 'site' in header_name:
                    cell.font = site_font
                    cell.alignment = align_center
                elif any(k in header_name for k in ['hrs', 'min', 'count', 'penalty', 'inr']):
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left

        # Auto-adjust Column Widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(out)
    return out.getvalue()

with tab_h:
    if h_sites.empty:
        st.info("HCIN site data nahi")
    else:
        st.metric("Unique sites (HCIN)", len(h_sites))
        st.dataframe(h_sites, use_container_width=True, height=420)
        st.download_button("📥 Download HCIN Site Summary (CSV)", h_sites.to_csv(index=False).encode('utf-8'),
                           file_name=f"Penalty_Sites_HCIN_{period.replace(' ','_')}.csv", mime="text/csv", key="h_sites_dl")

with tab_o:
    if o_sites.empty:
        st.info("ONEOTT site data nahi")
    else:
        st.metric("Unique sites (ONEOTT)", len(o_sites))
        st.dataframe(o_sites, use_container_width=True, height=420)
        st.download_button("📥 Download ONEOTT Site Summary (CSV)", o_sites.to_csv(index=False).encode('utf-8'),
                           file_name=f"Penalty_Sites_OTT_{period.replace(' ','_')}.csv", mime="text/csv", key="o_sites_dl")

with tab_all:
    if not h_sites.empty:
        h_sites_copy = h_sites.copy()
        h_sites_copy.insert(0, 'ISP', 'HCIN')
    else:
        h_sites_copy = pd.DataFrame()

    if not o_sites.empty:
        o_sites_copy = o_sites.copy()
        o_sites_copy.insert(0, 'ISP', 'ONEOTT')
    else:
        o_sites_copy = pd.DataFrame()

    combined = pd.concat([h_sites_copy, o_sites_copy], ignore_index=True) if not h_sites_copy.empty or not o_sites_copy.empty else pd.DataFrame()
    
    if combined.empty:
        st.info("No data available to display")
    else:
        st.dataframe(combined, use_container_width=True, height=420)

        st.download_button(
            "📊 Download Colorful Formatted Excel Report",
            data=get_styled_excel_bytes(h_sites, o_sites, combined),
            file_name=f"XTRNATE_Formatted_Penalty_Report_{period.replace(' ','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

st.markdown("---")
tab1, tab2 = st.tabs(["HCIN Breach Tickets", "ONEOTT Breach Tickets"])
show_cols = ['ticket_id', 'site_code', 'submitted_time', 'resolved_time', 'resolution_hours',
             'sla_status', 'penalty_est', 'state', 'reason_clean', 'owner']

with tab1:
    if not hcin.empty:
        hb = hcin[hcin['penalty_est'] > 0].sort_values('resolution_hours', ascending=False)
        if hb.empty:
            st.success("No HCIN breaches")
        else:
            cols = [c for c in show_cols if c in hb.columns]
            st.dataframe(hb[cols], use_container_width=True, height=350)

with tab2:
    if not ott.empty:
        ob = ott[ott['penalty_est'] > 0].sort_values('resolution_hours', ascending=False)
        if ob.empty:
            st.success("No ONEOTT breaches")
        else:
            cols = [c for c in show_cols if c in ob.columns]
            st.dataframe(ob[cols], use_container_width=True, height=350)
